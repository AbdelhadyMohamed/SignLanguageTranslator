from openpyxl import load_workbook
import os

# Defining constants
TRAINING_DATA_PATH = 'C:/Users/Abdelhady/Downloads/Arabic/ArabicDataset/ArASL_Database_54K'
EVALUATION_DATA_PATH = 'C:/Users/Abdelhady/Downloads/Arabic/ArabicDataset/ImagesToPred'
PARALLELIZM = 15
BATCH_SIZE = 32

# Reading dataset
label_map = {}

wb = load_workbook('C:/Users/Abdelhady/Downloads/Arabic/ArabicDataset/Labels/ClassLabels.xlsx')
ws = wb.active

it = ws.iter_rows()
next(it)
for row in it:
    label_map[row[1].value] = row[0].value

image_names = []
labels = []
for root, dirs, files in os.walk(TRAINING_DATA_PATH):

    for file in files:
        image_names.append(f'{root}/{file}')
        labels.append(label_map[os.path.basename(root)])
from keras.layers import Rescaling, RandomFlip, RandomZoom, RandomRotation, Resizing
from sklearn.model_selection import train_test_split
from keras import Sequential
import tensorflow as tf
import numpy as np
import cv2

# Splitting the data to train and test portions
train_images, test_images, train_labels, test_labels = train_test_split(image_names, labels, test_size=0.2,
                                                                        random_state=4)


# Creating a generator for our dataset
def dataset_generator(image_names, labels):
    image_names = [n.decode('utf-8') if type(n) == bytes else n for n in image_names]
    image_rescaler = Sequential([Rescaling(scale=1.0 / 255), Resizing(64, 64)])

    for i in range(len(image_names)):
        image = cv2.imread(os.path.join(TRAINING_DATA_PATH, image_names[i]))
        image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        image = np.expand_dims(image, axis=2)
        yield image_rescaler(image), labels[i]


# Creating an augmenter for our images
image_augmenter = Sequential([RandomFlip("horizontal"),
                              RandomZoom(height_factor=(-0.05, -0.15), width_factor=(-0.05, -0.15)),
                              RandomRotation(0.1)])

# Creating training and testing tensorflow objects using the dataset generator (for maximum speed & memory efficient training)
training_dataset = tf.data.Dataset.from_generator(dataset_generator,
                                                  args=(train_images, train_labels),

                                                  output_signature=(tf.TensorSpec(shape=(64, 64, 1), dtype=tf.float16),
                                                                    tf.TensorSpec(shape=(), dtype=tf.uint8)))
testing_dataset = tf.data.Dataset.from_generator(dataset_generator,
                                                 args=(test_images, test_labels),
                                                 output_signature=(tf.TensorSpec(shape=(64, 64, 1), dtype=tf.float16),
                                                                   tf.TensorSpec(shape=(), dtype=tf.uint8)))

# Configuring our datasets for the training process
training_dataset = training_dataset.map(lambda x, y: [image_augmenter(x, training=True), y],
                                        num_parallel_calls=PARALLELIZM).batch(BATCH_SIZE).prefetch(PARALLELIZM)
testing_dataset = testing_dataset.batch(BATCH_SIZE).prefetch(PARALLELIZM)
from keras.layers import Conv2D, MaxPooling2D, Flatten, Dense
from tensorflow.keras.models import load_model
from keras.optimizers import adam_v2
import matplotlib.pyplot as plt
from keras import Sequential
import os

# Constructing the model
model = Sequential()

model.add(Conv2D(filters=32, kernel_size=(3, 3), activation="relu", input_shape=(64, 64, 1)))
model.add(MaxPooling2D(2, 2, padding='same'))

model.add(Conv2D(filters=128, kernel_size=(3, 3), activation="relu"))
model.add(MaxPooling2D(2, 2, padding='same'))

model.add(Conv2D(filters=512, kernel_size=(3, 3), activation="relu"))
model.add(MaxPooling2D(2, 2, padding='same'))

model.add(Flatten())
model.add(Dense(units=1024, activation="relu"))
model.add(Dense(units=256, activation="relu"))
model.add(Dense(units=32, activation="softmax"))

model.summary()

# Training the model
model.compile(optimizer=adam_v2.Adam(),
              loss=tf.keras.losses.sparse_categorical_crossentropy,
              metrics=['accuracy'])

history = model.fit(training_dataset, epochs=20, validation_data=testing_dataset)

plt.plot(history.history['accuracy'], label='accuracy')
plt.plot(history.history['val_accuracy'], label='val_accuracy')
plt.xlabel('Epoch')
plt.ylabel('Accuracy')
plt.legend(loc='lower right')
plt.show()

model.save('Model.h5')
model.save_weights('/Weights/weights')
from sklearn.metrics import confusion_matrix
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import tensorflow as tf

# ... Existing code ...

# Load the trained model
model = tf.keras.models.load_model('Model.h5')

# ... Existing code ...

# Calculate the confusion matrix
true_labels = []
predicted_labels = []

for batch in testing_dataset:
    images, labels = batch
    predictions = model.predict(images)
    batch_predicted_labels = np.argmax(predictions, axis=1)
    predicted_labels.extend(batch_predicted_labels)
    true_labels.extend(labels.numpy())

confusion_mat = confusion_matrix(true_labels, predicted_labels)

# Visualize the confusion matrix
labels = sorted(set(true_labels))
sns.heatmap(confusion_mat, annot=True, fmt='d', xticklabels=labels, yticklabels=labels)
plt.xlabel('Predicted')
plt.ylabel('True')
plt.show()
from sklearn.metrics import confusion_matrix, classification_report
from tensorflow.keras.models import load_model

# Load the trained model
model = load_model('Model.h5')

# Generate predictions on the testing dataset
y_pred = model.predict(testing_dataset)
y_pred = np.argmax(y_pred, axis=1)

# Generate the confusion matrix
cm = confusion_matrix(test_labels, y_pred)
print("Confusion Matrix:")
print(cm)

# Generate the classification report
report = classification_report(test_labels, y_pred)
print("Classification Report:")
print(report)
import matplotlib.pyplot as plt
import seaborn as sns

# Generate the confusion matrix
cm = confusion_matrix(test_labels, y_pred)

# Plot the confusion matrix
plt.figure(figsize=(12, 12))
sns.heatmap(cm, annot=True, fmt="d", cmap="Blues", cbar=False)
plt.title("Confusion Matrix")
plt.xlabel("Predicted Labels")
plt.ylabel("True Labels")
plt.show()
